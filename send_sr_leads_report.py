"""
send_sr_leads_report.py: adapted from send_leads_report.py

Generates an .xlsx of CONFIRMED (+ optionally UNCERTAIN) SR007 leads and
emails it EVERY run; unlike the YC version, which stays silent when
nothing changed. Here, silence is itself ambiguous ("did the pipeline even
run today, or is there really nothing new?"), so instead:
  - Nothing changed: send a short "no updates since <date>" email anyway
  - Something changed: send "N new lead(s): Name1, Name2, ..."
Both cases still attach the current spreadsheet.

Recipients default to whatever's in sr_config.json (edit via
`python sr_configure.py`, not by hand); --recipients overrides that for
a single run without touching the saved config.

SMTP config matches send_leads_report.py exactly (Gmail App Password
approach; see README.md from the original YC pipeline):
    export SMTP_USERNAME=your_sending_address@gmail.com
    export SMTP_PASSWORD=your_16_char_app_password
"""

import argparse
import json
import smtplib
import ssl
import sys
from datetime import datetime, timezone
from email.message import EmailMessage
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment

from sr_config import load_config
from sr_queries import CURRENT_BATCH

# Namespaced by batch: see the comment on ROSTER_PATH in sr_unified_leads.py.
CLASSIFIED_PATH = Path(f"sr_roster_classified_{CURRENT_BATCH}.json")
LAST_SENT_PATH = Path(f"sr_last_sent_{CURRENT_BATCH}.json")
OUT_XLSX = Path(f"{CURRENT_BATCH.lower()}_leads_report.xlsx")

SMTP_HOST = "smtp.gmail.com"
SMTP_PORT = 465


def human_date(iso_str: str) -> str:
    """'2026-08-25T14:40:03+00:00' -> 'Aug 25, 2026'; the ISO timestamps
    used everywhere else in this pipeline are precise but not what anyone
    wants to read in an email or a spreadsheet cell."""
    dt = datetime.fromisoformat(iso_str)
    return f"{dt:%b} {dt.day}, {dt:%Y}"


def _require_env(name: str) -> str:
    import os
    value = os.environ.get(name)
    if not value:
        raise RuntimeError(f"Missing required environment variable '{name}'.")
    return value


def build_xlsx(classified):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{CURRENT_BATCH} Leads"

    headers = ["Company", "Verdict", "Reason", "First Seen", "Sources", "Best Evidence"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Arial", bold=True)

    row = 2
    ordered = sorted(
        classified.items(),
        key=lambda kv: {"CONFIRMED": 0, "UNCERTAIN": 1, "REJECTED": 2}[kv[1]["verdict"]],
    )
    for _, entry in ordered:
        if entry["verdict"] == "REJECTED":
            continue
        sources = ", ".join(sorted({e["source"] for e in entry["evidence"]}))
        best_evidence = entry["evidence"][0]["snippet"] if entry["evidence"] else ""
        ws.cell(row=row, column=1, value=entry["display_name"]).font = Font(name="Arial")
        ws.cell(row=row, column=2, value=entry["verdict"]).font = Font(name="Arial")
        ws.cell(row=row, column=3, value=entry["verdict_reason"]).font = Font(name="Arial")
        ws.cell(row=row, column=4, value=human_date(entry["first_seen"])).font = Font(name="Arial")
        ws.cell(row=row, column=5, value=sources).font = Font(name="Arial")
        c = ws.cell(row=row, column=6, value=best_evidence)
        c.font = Font(name="Arial")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    widths = [22, 12, 40, 12, 22, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(OUT_XLSX)
    return row - 2  # number of data rows written


def diff_against_last_sent(classified):
    """
    Returns (last_sent_at_iso_or_None, list_of_newly_confirmed_display_names).
    last_sent_at is None on a genuine first-ever run (no prior snapshot at
    all); distinct from "ran before, nothing new," which returns a date
    and an empty list.
    """
    current = {
        k: v["display_name"] for k, v in classified.items() if v["verdict"] == "CONFIRMED"
    }
    if not LAST_SENT_PATH.exists():
        return None, list(current.values())

    last = json.loads(LAST_SENT_PATH.read_text())
    last_confirmed_keys = set(last.get("confirmed", []))
    new_names = [name for k, name in current.items() if k not in last_confirmed_keys]
    return last.get("sent_at"), new_names


def send_email(recipients: list[str], subject: str, body: str):
    smtp_username = _require_env("SMTP_USERNAME")
    smtp_password = _require_env("SMTP_PASSWORD")

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = smtp_username
    msg["To"] = ", ".join(recipients)
    msg.set_content(body)

    with open(OUT_XLSX, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=OUT_XLSX.name,
        )

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=context) as s:
        s.login(smtp_username, smtp_password)
        s.send_message(msg)


def run(recipients: list[str]):
    if not CLASSIFIED_PATH.exists():
        print(f"ERROR: {CLASSIFIED_PATH} not found. Run classify_sr_leads.py first.", file=sys.stderr)
        sys.exit(1)

    classified = json.loads(CLASSIFIED_PATH.read_text())
    n_rows = build_xlsx(classified)
    print(f"Wrote spreadsheet: {OUT_XLSX}")

    last_sent_at, new_names = diff_against_last_sent(classified)
    today = human_date(datetime.now(timezone.utc).isoformat())
    confirmed = [k for k, v in classified.items() if v["verdict"] == "CONFIRMED"]

    if new_names:
        subject = f"[a16z Speedrun {CURRENT_BATCH}] {len(new_names)} new lead(s): {', '.join(new_names)}"
        body_lines = [f"New confirmed {CURRENT_BATCH} lead(s) as of {today}:", ""]
        body_lines += [f"  - {name}" for name in new_names]
        body_lines += ["", f"Full current list ({len(confirmed)} confirmed) attached."]
    else:
        since = human_date(last_sent_at) if last_sent_at else today
        subject = f"[a16z Speedrun {CURRENT_BATCH}] No updates since {since}"
        body_lines = [
            f"No new confirmed leads since the last check ({since}).",
            f"Still tracking {len(confirmed)} confirmed compan{'y' if len(confirmed) == 1 else 'ies'}.",
            "Full current list attached.",
        ]

    try:
        send_email(recipients, subject, "\n".join(body_lines))
        print(f"Email sent to {', '.join(recipients)}: {subject}")
    except Exception as e:
        print(f"ERROR sending email: {e}", file=sys.stderr)
        print("Spreadsheet was still written; fix email config and re-run to send the alert.", file=sys.stderr)
        return

    LAST_SENT_PATH.write_text(json.dumps({"confirmed": confirmed, "sent_at": datetime.now(timezone.utc).isoformat()}))
    print(f"Report sent: {n_rows} leads ({len(confirmed)} confirmed).")


def main():
    config = load_config()
    parser = argparse.ArgumentParser(description="Write SR007 xlsx + email it (every run)")
    parser.add_argument(
        "--recipients",
        nargs="+",
        default=config["recipients"],
        help="Override the configured recipient list for this run only (space-separated addresses)",
    )
    args = parser.parse_args()
    run(args.recipients)


if __name__ == "__main__":
    main()
