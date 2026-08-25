"""
send_leads_report.py

Takes the leads JSON produced by li_unified_leads.py (optionally enriched
by classify_leads.py) and:
  1. Writes a real spreadsheet (leads_<batch>.xlsx); the deliverable
     replaces the earlier HTML dashboard per request.
  2. Emails querovaleria04@gmail.com:
       - ALWAYS on the very first run (baseline)
       - ONLY on later runs if the set of leads actually changed since
         the last run
     Otherwise stays silent: no email every single run regardless of
     content, which is the whole point of the diff check.

This script does NOT do any scraping itself; run li_unified_leads.py
(and classify_leads.py, if you want AI relevance filtering included)
first, then run this on its output. Keeping this separate means this
piece can be scheduled/repeated independently of the scraping steps.

Setup (reuses the same Gmail App Password approach from earlier in this
project; see README.md from the very first version of this pipeline):
    pip install openpyxl
    export SMTP_USERNAME=your_sending_address@gmail.com
    export SMTP_PASSWORD=your_16_char_app_password

Usage:
    python send_leads_report.py --input leads_fall-2026.json
"""

import argparse
import json
import smtplib
import ssl
import sys
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

SMTP_HOST = "smtp.gmail.com"
SMTP_PORT = 465

SNAPSHOT_DIR = Path("report_snapshots")

COLUMNS = [
    "Name", "Company (guess)", "Status", "YC Match", "Sources",
    "AI Relevant", "AI Reason", "Link",
]


def _require_env(name: str) -> str:
    import os
    value = os.environ.get(name)
    if not value:
        raise RuntimeError(f"Missing required environment variable '{name}'.")
    return value


def write_spreadsheet(data: dict, out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_fill = PatternFill(start_color="1D1A15", end_color="1D1A15", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill

    row_idx = 2
    for lead in data.get("leads", []):
        status = "Unconfirmed" if lead.get("pre_mfn") else "Confirmed"
        values = [
            lead.get("name") or "(name unknown)",
            lead.get("company_guess") or "",
            status,
            lead.get("yc_match") or "",
            ", ".join(lead.get("sources", [])),
            "" if lead.get("ai_relevant") is None else ("Yes" if lead["ai_relevant"] else "No"),
            lead.get("ai_reason") or "",
            lead.get("link") or "",
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        # Light highlight for unconfirmed (pre-MFN) rows: the leads that matter most
        if lead.get("pre_mfn"):
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color="FFF6D6", end_color="FFF6D6", fill_type="solid"
                )
        row_idx += 1

    for col_idx, name in enumerate(COLUMNS, start=1):
        # Rough auto-width based on header/content length
        max_len = len(name)
        for row in range(2, row_idx):
            v = ws.cell(row=row, column=col_idx).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"
    wb.save(out_path)


def snapshot_path(batch_slug: str) -> Path:
    SNAPSHOT_DIR.mkdir(exist_ok=True)
    return SNAPSHOT_DIR / f"{batch_slug}.json"


def lead_signature(lead: dict) -> str:
    """
    A stable identifier per lead for change detection. Uses the link
    (unique per person/post) rather than the whole lead dict, so a
    cosmetic change (e.g. sources list gaining an extra tag on a re-scan)
    doesn't itself count as "new"; only a genuinely new lead does.
    """
    return lead.get("link", "") or lead.get("name", "")


def load_last_snapshot(batch_slug: str) -> set:
    path = snapshot_path(batch_slug)
    if not path.exists():
        return None  # None = no prior run at all, distinct from an empty set
    try:
        return set(json.loads(path.read_text()).get("signatures", []))
    except (json.JSONDecodeError, OSError):
        return None


def save_snapshot(batch_slug: str, signatures: set):
    snapshot_path(batch_slug).write_text(json.dumps({"signatures": sorted(signatures)}, indent=2))


def build_email(data: dict, is_first_run: bool, new_leads: list, xlsx_path: Path, recipients: list[str]) -> MIMEMultipart:
    leads = data.get("leads", [])
    pre_mfn_count = sum(1 for l in leads if l.get("pre_mfn"))
    batch_tag = data.get("batch_tag", "")

    msg = MIMEMultipart()
    msg["From"] = _require_env("SMTP_USERNAME")
    msg["To"] = ", ".join(recipients)

    if is_first_run:
        msg["Subject"] = f"[MFC Sourcing] Baseline: {len(leads)} {batch_tag} leads ({pre_mfn_count} pre-MFN)"
        body_lines = [
            f"First run for {batch_tag}: {len(leads)} total leads found, "
            f"{pre_mfn_count} not yet on YC's public directory.",
            "",
            "Full list attached as a spreadsheet.",
            "",
            "From here on, you'll only get an email when the list actually changes.",
        ]
    else:
        msg["Subject"] = f"[MFC Sourcing] {len(new_leads)} new {batch_tag} lead(s)"
        body_lines = [f"{len(new_leads)} new lead(s) since the last check:", ""]
        for lead in new_leads:
            body_lines.append(f"  - {lead.get('name') or '(name unknown)'}")
            if lead.get("company_guess"):
                body_lines.append(f"      Company: {lead['company_guess']}")
            body_lines.append(f"      {lead.get('link', '')}")
            body_lines.append("")
        body_lines.append("Full updated list attached as a spreadsheet.")

    msg.attach(MIMEText("\n".join(body_lines), "plain"))

    with open(xlsx_path, "rb") as f:
        part = MIMEApplication(f.read(), Name=xlsx_path.name)
    part["Content-Disposition"] = f'attachment; filename="{xlsx_path.name}"'
    msg.attach(part)

    return msg


def send_email(msg: MIMEMultipart):
    smtp_username = _require_env("SMTP_USERNAME")
    smtp_password = _require_env("SMTP_PASSWORD")
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=context) as server:
        server.login(smtp_username, smtp_password)
        server.send_message(msg)


def run(input_path: Path, recipients: list[str]):
    """
    Factored out of main() so yc_run_pipeline.py can call this directly
    instead of shelling out.
    """
    if not input_path.exists():
        print(f"ERROR: {input_path} not found. Run li_unified_leads.py first.", file=sys.stderr)
        sys.exit(1)

    data = json.loads(input_path.read_text())
    batch_slug = data.get("yc_batch_slug", "batch")
    leads = data.get("leads", [])

    xlsx_path = Path(f"leads_{batch_slug}.xlsx")
    write_spreadsheet(data, xlsx_path)
    print(f"Wrote spreadsheet: {xlsx_path}")

    current_signatures = {lead_signature(l) for l in leads}
    last_signatures = load_last_snapshot(batch_slug)

    is_first_run = last_signatures is None
    new_leads = []
    if not is_first_run:
        new_sigs = current_signatures - last_signatures
        new_leads = [l for l in leads if lead_signature(l) in new_sigs]

    should_email = is_first_run or bool(new_leads)

    if should_email:
        try:
            msg = build_email(data, is_first_run, new_leads, xlsx_path, recipients)
            send_email(msg)
            print(f"Email sent to {', '.join(recipients)} ({'baseline' if is_first_run else f'{len(new_leads)} new lead(s)'}).")
        except Exception as e:
            print(f"ERROR sending email: {e}", file=sys.stderr)
            print("Spreadsheet was still written; fix email config and re-run to send the alert.", file=sys.stderr)
    else:
        print("No change since last run; no email sent.")

    save_snapshot(batch_slug, current_signatures)


def main():
    parser = argparse.ArgumentParser(description="Write spreadsheet + conditionally email leads")
    parser.add_argument("--input", default="leads_fall-2026.json")
    parser.add_argument(
        "--recipients",
        nargs="+",
        default=None,
        help="Override the configured recipient list for this run only (space-separated addresses)",
    )
    args = parser.parse_args()

    recipients = args.recipients
    if recipients is None:
        from yc_config import load_config
        recipients = load_config()["recipients"]

    run(Path(args.input), recipients)


if __name__ == "__main__":
    main()
